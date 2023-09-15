#!/var/www/html/SPIRAL.web.tool/spiral_venv/bin/python3.10

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import string
from SPIRAL_basic_funcs import *


####################################################################################################################
def design_excel(sigfile):
    print('design_excel!')

    # start from pyxl
    wb = openpyxl.load_workbook(sigfile)
    ws = wb.active

    if ws.max_row > 1:
        col_list = list(string.ascii_uppercase) + ['A' + letter for letter in list(string.ascii_uppercase)] + [
            'B' + letter for letter in list(string.ascii_uppercase)]
        for col in col_list[1:ws.max_column]:
            if ws[col + '1'].value in ['proc_link', 'func_link', 'comp_link', 'networks', 'UMAPs', 'PCAs',
                                       'orig_UMAPs', 'orig_PCAs', 'spatial', 'Gnetworks', 'GUMAPs', 'GPCAs',
                                       'GUMAPs_orig', 'GPCAs_orig', 'Gspatial',
                                       'repcells_UMAPs', 'repcells_PCAs', 'repcells_Gumaps',
                                       'repcells_GPCAs', 'orig_GUMAPs', 'orig_GPCAs',
                                       'spots_UMAPs', 'spots_PCAs', 'spots_Gumaps',
                                       'spots_GPCAs', 'samples_UMAPs', 'samples_PCAs', 'samples_Gumaps',
                                       'samples_GPCAs']:
                # add underline and blue color
                for row in np.arange(2, ws.max_row + 1):
                    cell = ws[col + str(row)]
                    cell.font = Font(color='0000FF', underline='single')
            elif ws[col + '1'].value == 'num_genes_in_struct':
                # Add a three-color scale
                ws.conditional_formatting.add(col + '2:' + col + str(ws.max_row),
                                              ColorScaleRule(start_type='percentile', start_value=10,
                                                             start_color='FFFFFF',
                                                             mid_type='percentile', mid_value=50, mid_color='FF8080',
                                                             end_type='percentile', end_value=90, end_color='FF0000'))
            elif ws[col + '1'].value in ['num_cells_in_struct', 'num_samples_in_struct', 'num_spots_in_struct']:
                ws.conditional_formatting.add(col + '2:' + col + str(ws.max_row),
                                              ColorScaleRule(start_type='percentile', start_value=10,
                                                             start_color='FFFFFF',
                                                             mid_type='percentile', mid_value=50, mid_color='FF8080',
                                                             end_type='percentile', end_value=90, end_color='FF0000'))
            elif ws[col + '1'].value == 'structure_average_std':
                ws.conditional_formatting.add(col + '2:' + col + str(ws.max_row),
                                              ColorScaleRule(start_type='percentile', start_value=10,
                                                             start_color='00FF00',
                                                             mid_type='percentile', mid_value=50, mid_color='FF6600',
                                                             end_type='percentile', end_value=90, end_color='FF0000'))
            elif ws[col + '1'].value == "log10_pavg_of_genes":
                # Add fill
                for row in np.arange(2, ws.max_row + 1):
                    cell = ws[col + str(row)]
                    cell.fill = PatternFill("solid", fgColor="00FFFF")
            elif ws[col + '1'].value == 'log10_corrected_pval':
                # Add fill
                for row in np.arange(2, ws.max_row + 1):
                    cell = ws[col + str(row)]
                    cell.fill = PatternFill("solid", fgColor="99CC00")

    wb.save(sigfile)
